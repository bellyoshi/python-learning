import openpyxl
import os
from pathlib import Path

def read_excel_file(file_path):
    """
    エクセルファイルの中身を確認する関数
    """
    try:
        # ワークブックを開く
        workbook = openpyxl.load_workbook(file_path)
        
        print(f"=== エクセルファイル情報 ===")
        print(f"ファイル名: {os.path.basename(file_path)}")
        print(f"シート数: {len(workbook.sheetnames)}")
        print(f"シート名: {workbook.sheetnames}")
        print()
        
        # 各シートの内容を確認
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"=== シート: {sheet_name} ===")
            print(f"最大行: {sheet.max_row}")
            print(f"最大列: {sheet.max_column}")
            print()
            
            # データの内容を表示（最初の10行まで）
            print("データ内容（最初の10行）:")
            for row in range(1, min(11, sheet.max_row + 1)):
                row_data = []
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                print(f"行{row}: {row_data}")
            print()
            
            # セルの値の型も確認
            print("セルの値の型（最初の5行×5列）:")
            for row in range(1, min(6, sheet.max_row + 1)):
                for col in range(1, min(6, sheet.max_column + 1)):
                    cell = sheet.cell(row=row, column=col)
                    cell_value = cell.value
                    cell_type = type(cell_value).__name__ if cell_value is not None else "None"
                    print(f"({row},{col}): {cell_value} (型: {cell_type})")
            print("-" * 50)
            print()
        
        workbook.close()
        
    except FileNotFoundError:
        print(f"エラー: ファイル '{file_path}' が見つかりません。")
    except Exception as e:
        print(f"エラー: {e}")

def main():
    """
    メイン関数
    """
    # 現在のディレクトリにあるエクセルファイルを探す
    excel_files = list(Path('.').glob('*.xlsx'))
    
    if not excel_files:
        print("エクセルファイル（.xlsx）が見つかりません。")
        return
    
    print(f"見つかったエクセルファイル: {[f.name for f in excel_files]}")
    print()
    
    # 各エクセルファイルを読み込む
    for excel_file in excel_files:
        print(f"読み込み中: {excel_file}")
        read_excel_file(excel_file)
        print("=" * 60)
        print()

if __name__ == "__main__":
    main() 