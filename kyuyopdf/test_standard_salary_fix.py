#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
標準報酬月額の修正テストプログラム
"""

import openpyxl
import os

def test_standard_salary_fix():
    """標準報酬月額の修正をテスト"""
    excel_file = "給与支給一覧令和7年.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"エラー: {excel_file} が見つかりません")
        return
    
    print("=== 標準報酬月額の修正テスト ===")
    print(f"ファイル: {excel_file}")
    print()
    
    # データのみの読み込み
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb.active
    
    # 渡邉俊行さんの1月のデータを確認
    print("=== 渡邉俊行さんの1月データ確認 ===")
    for row in range(2, 100):
        name_cell = ws.cell(row=row, column=3).value  # 氏名列（3列目）
        date_cell = ws.cell(row=row, column=1).value  # 支給日列（1列目）
        
        if name_cell == "渡邉俊行" and date_cell and hasattr(date_cell, 'month') and date_cell.month == 1:
            print(f"渡邉俊行さんの1月データ（行{row}）:")
            
            # 主要な項目を確認
            items = [
                ('総支給額', 4),
                ('標準報酬月額', 5),
                ('健康保険', 6),
                ('厚生年金', 7),
                ('社会保険料控除後', 11),
                ('源泉所得税', 13),
                ('差引支給額', 14),
                ('振込金額', 21)
            ]
            
            for item_name, col in items:
                value = ws.cell(row=row, column=col).value
                print(f"  {item_name}: {value}")
            
            break
    
    wb.close()

if __name__ == "__main__":
    test_standard_salary_fix() 