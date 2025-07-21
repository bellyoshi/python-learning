#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
個人給与データクラス
ExcelColumnsで定義されているプロパティのみを含む
"""

from dataclasses import dataclass
from datetime import datetime
from typing import Optional, List, Dict, Any
from decimal import Decimal
from excel_columns import ExcelColumns
import openpyxl
from openpyxl.utils import get_column_letter

@dataclass
class PersonalSalaryData:
    """個人給与データクラス - ExcelColumnsのプロパティのみを含む"""
    
    # 給与データシートのプロパティ
    支給日: Optional[datetime] = None
    氏名: Optional[str] = None
    総支給額: Optional[Decimal] = None
    標準報酬月額: Optional[Decimal] = None
    健康保険: Optional[Decimal] = None
    厚生年金: Optional[Decimal] = None
    社会保険料控除後: Optional[Decimal] = None
    源泉所得税: Optional[Decimal] = None
    差引支給額: Optional[Decimal] = None
    振込金額: Optional[Decimal] = None
    健康保険料_従業員: Optional[Decimal] = None
    厚生年金_従業員: Optional[Decimal] = None
    社会保険料控除額: Optional[Decimal] = None
    健康保険料_会社負担: Optional[Decimal] = None
    厚生年金_会社負担: Optional[Decimal] = None
    賃料控除: Optional[Decimal] = None
    駐車場控除: Optional[Decimal] = None
    扶養親族等の数: Optional[int] = None
    
    @classmethod
    def from_excel_row(cls, row_data: List[Any], sheet_name: str = "給与データ") -> 'PersonalSalaryData':
        """
        エクセルの行データからPersonalSalaryDataインスタンスを作成
        
        Args:
            row_data: エクセルの行データ（リスト）
            sheet_name: シート名（デフォルト: "給与データ"）
            
        Returns:
            PersonalSalaryDataインスタンス
        """
        # 列番号からデータを取得（1ベースのインデックス）
        def get_cell_value(column_index: int) -> Any:
            """指定された列の値を取得（1ベースのインデックス）"""
            if column_index <= len(row_data):
                return row_data[column_index - 1]
            return None
        
        def parse_decimal(value: Any) -> Optional[Decimal]:
            """値をDecimal型に変換"""
            if value is None or value == '':
                return None
            try:
                if isinstance(value, str):
                    # カンマを除去してから変換
                    value = value.replace(',', '')
                return Decimal(str(value))
            except (ValueError, TypeError):
                return None
        
        def parse_datetime(value: Any) -> Optional[datetime]:
            """値をdatetime型に変換"""
            if value is None or value == '':
                return None
            try:
                if isinstance(value, datetime):
                    return value
                # openpyxlの日付オブジェクトの場合
                if hasattr(value, 'date'):
                    return value
                return None
            except (ValueError, TypeError):
                return None
        
        def parse_int(value: Any) -> Optional[int]:
            """値をint型に変換"""
            if value is None or value == '':
                return None
            try:
                return int(value)
            except (ValueError, TypeError):
                return None
        
        # ExcelColumnsの定義に基づいてデータを取得
        personal_data = cls()
        
        # 支給日
        payment_date_value = get_cell_value(ExcelColumns.PAYMENT_DATE)
        personal_data.支給日 = parse_datetime(payment_date_value)
        
        # 氏名
        personal_data.氏名 = get_cell_value(ExcelColumns.EMPLOYEE_NAME)
        
        # 総支給額
        total_salary_value = get_cell_value(ExcelColumns.TOTAL_SALARY)
        personal_data.総支給額 = parse_decimal(total_salary_value)
        
        # 標準報酬月額
        standard_reward_value = get_cell_value(ExcelColumns.STANDARD_REWARD)
        personal_data.標準報酬月額 = parse_decimal(standard_reward_value)
        
        # 健康保険
        health_insurance_value = get_cell_value(ExcelColumns.HEALTH_INSURANCE)
        personal_data.健康保険 = parse_decimal(health_insurance_value)
        
        # 厚生年金
        kousei_insurance_value = get_cell_value(ExcelColumns.KOUSEI_INSURANCE)
        personal_data.厚生年金 = parse_decimal(kousei_insurance_value)
        
        # 社会保険料控除後
        social_insurance_after_value = get_cell_value(ExcelColumns.SOCIAL_INSURANCE_AFTER)
        personal_data.社会保険料控除後 = parse_decimal(social_insurance_after_value)
        
        # 源泉所得税
        income_tax_value = get_cell_value(ExcelColumns.INCOME_TAX)
        personal_data.源泉所得税 = parse_decimal(income_tax_value)
        
        # 振込金額
        transfer_amount_value = get_cell_value(ExcelColumns.TRANSFER_AMOUNT)
        personal_data.振込金額 = parse_decimal(transfer_amount_value)
        
        # 健康保険料_従業員
        health_insurance_employee_value = get_cell_value(ExcelColumns.HEALTH_INSURANCE_EMPLOYEE)
        personal_data.健康保険料_従業員 = parse_decimal(health_insurance_employee_value)
        
        # 厚生年金_従業員
        pension_employee_value = get_cell_value(ExcelColumns.PENSION_EMPLOYEE)
        personal_data.厚生年金_従業員 = parse_decimal(pension_employee_value)
        
        # 社会保険料控除額
        social_insurance_deduction_value = get_cell_value(ExcelColumns.SOCIAL_INSURANCE_DEDUCTION)
        personal_data.社会保険料控除額 = parse_decimal(social_insurance_deduction_value)
        
        # 健康保険料_会社負担
        health_insurance_company_value = get_cell_value(ExcelColumns.HEALTH_INSURANCE_COMPANY)
        personal_data.健康保険料_会社負担 = parse_decimal(health_insurance_company_value)
        
        # 厚生年金_会社負担
        pension_company_value = get_cell_value(ExcelColumns.PENSION_COMPANY)
        personal_data.厚生年金_会社負担 = parse_decimal(pension_company_value)
        
        # 賃料控除
        rent_deduction_value = get_cell_value(ExcelColumns.RENT_DEDUCTION)
        personal_data.賃料控除 = parse_decimal(rent_deduction_value)
        
        # 駐車場控除
        parking_deduction_value = get_cell_value(ExcelColumns.PARKING_DEDUCTION)
        personal_data.駐車場控除 = parse_decimal(parking_deduction_value)
        
        # 扶養親族等の数
        dependents_count_value = get_cell_value(ExcelColumns.DEPENDENTS_COUNT)
        personal_data.扶養親族等の数 = parse_int(dependents_count_value)
        
        return personal_data
    
    @classmethod
    def from_excel_file(cls, file_path: str, sheet_name: str = "給与データ", 
                       row_number: Optional[int] = None, employee_name: Optional[str] = None) -> List['PersonalSalaryData']:
        """
        エクセルファイルからPersonalSalaryDataインスタンスのリストを作成
        
        Args:
            file_path: エクセルファイルのパス
            sheet_name: シート名（デフォルト: "給与データ"）
            row_number: 特定の行番号（1ベース、Noneの場合は全行）
            employee_name: 特定の従業員名（Noneの場合は全員）
            
        Returns:
            PersonalSalaryDataインスタンスのリスト
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook[sheet_name]
            
            personal_data_list = []
            
            # ヘッダー行をスキップ（2行目からデータ開始）
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_idx == 1:  # ヘッダー行をスキップ
                    continue
                
                # 空行をスキップ
                if not any(cell for cell in row):
                    continue
                
                # 特定の行番号が指定されている場合
                if row_number is not None and row_idx != row_number:
                    continue
                
                # 特定の従業員名が指定されている場合
                if employee_name is not None:
                    employee_name_cell = row[ExcelColumns.EMPLOYEE_NAME - 1] if len(row) >= ExcelColumns.EMPLOYEE_NAME else None
                    if employee_name_cell != employee_name:
                        continue
                
                # PersonalSalaryDataインスタンスを作成
                personal_data = cls.from_excel_row(list(row), sheet_name)
                personal_data_list.append(personal_data)
            
            workbook.close()
            return personal_data_list
            
        except Exception as e:
            print(f"エクセルファイルの読み込みエラー: {e}")
            return []
    
    def to_dict(self) -> Dict[str, Any]:
        """PersonalSalaryDataを辞書形式に変換"""
        return {
            '支給日': self.支給日.isoformat() if self.支給日 else None,
            '氏名': self.氏名,
            '総支給額': str(self.総支給額) if self.総支給額 is not None else None,
            '標準報酬月額': str(self.標準報酬月額) if self.標準報酬月額 is not None else None,
            '健康保険': str(self.健康保険) if self.健康保険 is not None else None,
            '厚生年金': str(self.厚生年金) if self.厚生年金 is not None else None,
            '社会保険料控除後': str(self.社会保険料控除後) if self.社会保険料控除後 is not None else None,
            '源泉所得税': str(self.源泉所得税) if self.源泉所得税 is not None else None,
            '差引支給額': str(self.差引支給額) if self.差引支給額 is not None else None,
            '振込金額': str(self.振込金額) if self.振込金額 is not None else None,
            '健康保険料_従業員': str(self.健康保険料_従業員) if self.健康保険料_従業員 is not None else None,
            '厚生年金_従業員': str(self.厚生年金_従業員) if self.厚生年金_従業員 is not None else None,
            '社会保険料控除額': str(self.社会保険料控除額) if self.社会保険料控除額 is not None else None,
            '健康保険料_会社負担': str(self.健康保険料_会社負担) if self.健康保険料_会社負担 is not None else None,
            '厚生年金_会社負担': str(self.厚生年金_会社負担) if self.厚生年金_会社負担 is not None else None,
            '賃料控除': str(self.賃料控除) if self.賃料控除 is not None else None,
            '駐車場控除': str(self.駐車場控除) if self.駐車場控除 is not None else None,
            '扶養親族等の数': self.扶養親族等の数
        }


class PersonalSalaryDataManager:
    """個人給与データの管理クラス"""
    
    def __init__(self):
        self.personal_data_list: List[PersonalSalaryData] = []
    
    def add_personal_data(self, personal_data: PersonalSalaryData) -> None:
        """個人給与データを追加"""
        self.personal_data_list.append(personal_data)
    
    def add_personal_data_list(self, personal_data_list: List[PersonalSalaryData]) -> None:
        """個人給与データのリストを追加"""
        self.personal_data_list.extend(personal_data_list)
    
    def get_all_personal_data(self) -> List[PersonalSalaryData]:
        """全ての個人給与データを取得"""
        return self.personal_data_list.copy()
    
    def get_personal_data_by_employee(self, employee_name: str) -> List[PersonalSalaryData]:
        """従業員名で個人給与データを取得"""
        return [data for data in self.personal_data_list if data.氏名 == employee_name]
    
    def get_personal_data_by_employee_and_month(self, employee_name: str, target_month: int) -> List[PersonalSalaryData]:
        """従業員名と月で個人給与データを取得"""
        return [
            data for data in self.personal_data_list 
            if data.氏名 == employee_name and data.支給日 and data.支給日.month == target_month
        ]
    
    def get_available_employees(self) -> List[str]:
        """利用可能な従業員名のリストを取得"""
        return list(set(data.氏名 for data in self.personal_data_list if data.氏名))
    
    def get_available_months(self) -> List[int]:
        """利用可能な月のリストを取得"""
        months = []
        for data in self.personal_data_list:
            if data.支給日:
                months.append(data.支給日.month)
        return sorted(list(set(months)))
    
    def get_total_salary_by_employee(self, employee_name: str) -> Optional[Decimal]:
        """従業員の総支給額合計を取得"""
        employee_data = self.get_personal_data_by_employee(employee_name)
        total = sum(data.総支給額 for data in employee_data if data.総支給額)
        return total if total > 0 else None
    
    def export_to_dict_list(self) -> List[Dict[str, Any]]:
        """全ての個人給与データを辞書のリスト形式でエクスポート"""
        return [data.to_dict() for data in self.personal_data_list]
    
    def load_from_excel_file(self, file_path: str, sheet_name: str = "給与データ") -> None:
        """
        エクセルファイルから個人給与データを読み込み
        
        Args:
            file_path: エクセルファイルのパス
            sheet_name: シート名（デフォルト: "給与データ"）
        """
        try:
            personal_data_list = PersonalSalaryData.from_excel_file(file_path, sheet_name)
            self.add_personal_data_list(personal_data_list)
            print(f"エクセルファイルから {len(personal_data_list)} 件の個人給与データを読み込みました。")
        except Exception as e:
            print(f"エクセルファイルの読み込みエラー: {e}")
    
    def load_from_excel_file_by_employee(self, file_path: str, employee_name: str, 
                                       sheet_name: str = "給与データ") -> None:
        """
        エクセルファイルから特定の従業員の個人給与データを読み込み
        
        Args:
            file_path: エクセルファイルのパス
            employee_name: 従業員名
            sheet_name: シート名（デフォルト: "給与データ"）
        """
        try:
            personal_data_list = PersonalSalaryData.from_excel_file(
                file_path, sheet_name, employee_name=employee_name
            )
            self.add_personal_data_list(personal_data_list)
            print(f"エクセルファイルから従業員 '{employee_name}' の {len(personal_data_list)} 件の個人給与データを読み込みました。")
        except Exception as e:
            print(f"エクセルファイルの読み込みエラー: {e}")
    
    def load_from_excel_file_by_row(self, file_path: str, row_number: int, 
                                   sheet_name: str = "給与データ") -> None:
        """
        エクセルファイルから特定の行の個人給与データを読み込み
        
        Args:
            file_path: エクセルファイルのパス
            row_number: 行番号（1ベース）
            sheet_name: シート名（デフォルト: "給与データ"）
        """
        try:
            personal_data_list = PersonalSalaryData.from_excel_file(
                file_path, sheet_name, row_number=row_number
            )
            self.add_personal_data_list(personal_data_list)
            print(f"エクセルファイルから行 {row_number} の {len(personal_data_list)} 件の個人給与データを読み込みました。")
        except Exception as e:
            print(f"エクセルファイルの読み込みエラー: {e}")
    
    def clear_all_data(self) -> None:
        """全ての個人給与データをクリア"""
        self.personal_data_list.clear()
        print("全ての個人給与データをクリアしました。")
    
    def get_statistics(self) -> Dict[str, Any]:
        """個人給与データの統計情報を取得"""
        if not self.personal_data_list:
            return {}
        
        total_employees = len(self.get_available_employees())
        total_records = len(self.personal_data_list)
        available_months = self.get_available_months()
        
        # 総支給額の統計
        total_salaries = [data.総支給額 for data in self.personal_data_list if data.総支給額]
        avg_salary = sum(total_salaries) / len(total_salaries) if total_salaries else 0
        max_salary = max(total_salaries) if total_salaries else 0
        min_salary = min(total_salaries) if total_salaries else 0
        
        return {
            '総従業員数': total_employees,
            '総レコード数': total_records,
            '利用可能な月': available_months,
            '平均総支給額': avg_salary,
            '最大総支給額': max_salary,
            '最小総支給額': min_salary
        }

