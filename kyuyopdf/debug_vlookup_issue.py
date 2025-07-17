#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VLOOKUP関数と月額表の問題を調査するプログラム
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os

def debug_vlookup_issue():
    """VLOOKUP関数と月額表の問題を調査"""
    excel_file = "給与支給一覧令和7年.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"エラー: {excel_file} が見つかりません")
        return
    
    print("=== VLOOKUP関数と月額表の問題調査 ===")
    print(f"ファイル: {excel_file}")
    print()
    
    # 通常の読み込み（数式あり）
    wb_normal = openpyxl.load_workbook(excel_file, data_only=False)
    ws_normal = wb_normal.active
    
    # データのみの読み込み（数式の計算結果）
    wb_data = openpyxl.load_workbook(excel_file, data_only=True)
    ws_data = wb_data.active
    
    # 名前付き範囲の調査
    print("=== 名前付き範囲の調査 ===")
    if hasattr(wb_normal.defined_names, 'definedName'):
        print("定義された名前付き範囲:")
        for name in wb_normal.defined_names.definedName:
            print(f"  {name.name}: {name.attr_text}")
    else:
        print("名前付き範囲は定義されていません")
    
    print()
    
    # 月額表に関連するセルを探す
    print("=== 月額表に関連するセルの調査 ===")
    for row in range(1, 100):
        for col in range(1, 50):
            cell_normal = ws_normal.cell(row=row, column=col)
            cell_data = ws_data.cell(row=row, column=col)
            
            cell_value_normal = cell_normal.value
            cell_value_data = cell_data.value
            
            # 月額表に関連するキーワードを含むセルを探す
            if cell_value_normal and isinstance(cell_value_normal, str):
                if '月額' in str(cell_value_normal) or '標準' in str(cell_value_normal):
                    print(f"行{row}列{col} ({get_column_letter(col)}{row}): {cell_value_normal}")
                    print(f"  データ読み込み値: {cell_value_data}")
                    print()
    
    # VLOOKUP関数を含むセルの詳細調査
    print("=== VLOOKUP関数を含むセルの詳細調査 ===")
    for row in range(1, 100):
        for col in range(1, 50):
            cell_normal = ws_normal.cell(row=row, column=col)
            cell_data = ws_data.cell(row=row, column=col)
            
            if cell_normal.data_type == 'f' and cell_normal.value:
                formula = str(cell_normal.value)
                if 'VLOOKUP' in formula:
                    print(f"行{row}列{col} ({get_column_letter(col)}{row}):")
                    print(f"  数式: {formula}")
                    print(f"  データ読み込み値: {cell_data.value}")
                    print(f"  データ型: {type(cell_data.value)}")
                    print()
    
    # 特定の行（46行目以降）の詳細調査
    print("=== 特定行の詳細調査（46行目以降） ===")
    for row in range(46, 55):
        print(f"\n--- 行{row}の詳細 ---")
        for col in range(1, 25):
            cell_normal = ws_normal.cell(row=row, column=col)
            cell_data = ws_data.cell(row=row, column=col)
            
            cell_value_normal = cell_normal.value
            cell_value_data = cell_data.value
            cell_formula = cell_normal.value if cell_normal.data_type == 'f' else None
            
            # 重要な列のみ表示
            if col in [1, 2, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 16, 17, 18]:
                print(f"列{col} ({get_column_letter(col)}):")
                print(f"  通常読み込み値: {cell_value_normal}")
                print(f"  データ読み込み値: {cell_value_data}")
                if cell_formula:
                    print(f"  数式: {cell_formula}")
                print()
    
    # 月額表の範囲を推測して調査
    print("=== 月額表の範囲推測調査 ===")
    
    # 一般的な月額表の場所を調査
    possible_ranges = [
        (1, 1, 50, 30),   # 左上
        (1, 30, 50, 60),  # 右上
        (50, 1, 100, 30), # 左下
        (50, 30, 100, 60) # 右下
    ]
    
    for start_row, start_col, end_row, end_col in possible_ranges:
        print(f"\n範囲 {start_row}:{start_col} から {end_row}:{end_col} の調査:")
        
        # この範囲で数値データを探す
        numeric_cells = []
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws_data.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, (int, float)):
                    numeric_cells.append((row, col, cell.value))
        
        if numeric_cells:
            print(f"  数値セル: {len(numeric_cells)}個")
            # 最初の10個を表示
            for row, col, value in numeric_cells[:10]:
                print(f"    {get_column_letter(col)}{row}: {value}")
        else:
            print("  数値セルは見つかりませんでした")
    
    wb_normal.close()
    wb_data.close()

if __name__ == "__main__":
    debug_vlookup_issue() 