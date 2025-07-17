#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
標準本州月額の空白問題を調査するためのデバッグプログラム
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os

def debug_standard_salary():
    """標準本州月額の空白問題を調査"""
    excel_file = "給与支給一覧令和7年.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"エラー: {excel_file} が見つかりません")
        return
    
    print("=== 標準本州月額の空白問題調査 ===")
    print(f"ファイル: {excel_file}")
    print()
    
    # 通常の読み込み（数式あり）
    print("1. 通常の読み込み（数式あり）:")
    wb_normal = openpyxl.load_workbook(excel_file, data_only=False)
    ws_normal = wb_normal.active
    
    # データのみの読み込み（数式の計算結果）
    print("2. データのみの読み込み（数式の計算結果）:")
    wb_data = openpyxl.load_workbook(excel_file, data_only=True)
    ws_data = wb_data.active
    
    # 全セルを調査
    print("\n=== 全セルの調査 ===")
    for row in range(1, 20):  # 最初の20行を調査
        for col in range(1, 20):  # 最初の20列を調査
            cell_normal = ws_normal.cell(row=row, column=col)
            cell_data = ws_data.cell(row=row, column=col)
            
            # 標準本州月額に関連しそうなセルを特定
            cell_value_normal = cell_normal.value
            cell_value_data = cell_data.value
            cell_formula = cell_normal.value if cell_normal.data_type == 'f' else None
            
            # 標準本州月額に関連するキーワードを含むセルを探す
            keywords = ['標準', '本州', '月額', '基本', '給与', '給料', '賃金']
            is_related = False
            
            if cell_value_normal:
                for keyword in keywords:
                    if keyword in str(cell_value_normal):
                        is_related = True
                        break
            
            if is_related or (cell_formula and '標準' in str(cell_formula)):
                print(f"行{row}列{col} ({get_column_letter(col)}{row}):")
                print(f"  通常読み込み値: {cell_value_normal}")
                print(f"  データ読み込み値: {cell_value_data}")
                print(f"  数式: {cell_formula}")
                print(f"  データ型: {cell_normal.data_type}")
                print()
    
    # 特定のセル範囲を詳しく調査
    print("\n=== 特定セル範囲の詳細調査 ===")
    
    # 給与データが含まれていそうな範囲を調査
    for row in range(1, 50):
        for col in range(1, 30):
            cell_normal = ws_normal.cell(row=row, column=col)
            cell_data = ws_data.cell(row=row, column=col)
            
            cell_value_normal = cell_normal.value
            cell_value_data = cell_data.value
            cell_formula = cell_normal.value if cell_normal.data_type == 'f' else None
            
            # 数値データで、通常読み込みとデータ読み込みで値が異なるセルを特定
            if (cell_value_normal is not None and 
                cell_value_data is not None and 
                cell_value_normal != cell_value_data and
                isinstance(cell_value_normal, (int, float)) or isinstance(cell_value_data, (int, float))):
                
                print(f"行{row}列{col} ({get_column_letter(col)}{row}):")
                print(f"  通常読み込み値: {cell_value_normal} (型: {type(cell_value_normal)})")
                print(f"  データ読み込み値: {cell_value_data} (型: {type(cell_value_data)})")
                print(f"  数式: {cell_formula}")
                print()
    
    # 空白セルの調査
    print("\n=== 空白セルの調査 ===")
    blank_cells = []
    for row in range(1, 50):
        for col in range(1, 30):
            cell_normal = ws_normal.cell(row=row, column=col)
            cell_data = ws_data.cell(row=row, column=col)
            
            # データ読み込みで空白だが、通常読み込みで数式があるセル
            if (cell_data.value is None or cell_data.value == '') and cell_normal.data_type == 'f':
                blank_cells.append((row, col, cell_normal.value))
    
    if blank_cells:
        print(f"データ読み込みで空白だが数式があるセル: {len(blank_cells)}個")
        for row, col, formula in blank_cells[:10]:  # 最初の10個を表示
            print(f"  行{row}列{col} ({get_column_letter(col)}{row}): {formula}")
    else:
        print("データ読み込みで空白の数式セルは見つかりませんでした")
    
    wb_normal.close()
    wb_data.close()

if __name__ == "__main__":
    debug_standard_salary() 