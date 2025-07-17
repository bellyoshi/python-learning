#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excelファイルのヘッダー行を確認するプログラム
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os

def check_excel_headers():
    """Excelファイルのヘッダー行を確認"""
    excel_file = "給与支給一覧令和7年.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"エラー: {excel_file} が見つかりません")
        return
    
    print("=== Excelファイルのヘッダー行確認 ===")
    print(f"ファイル: {excel_file}")
    print()
    
    # データのみの読み込み
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb.active
    
    # ヘッダー行を探す（最初の20行を調査）
    header_row = None
    for row in range(1, 21):
        for col in range(1, 30):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str):
                # 給与に関連するキーワードを含むセルを探す
                keywords = ['給与', '支給', '報酬', '標準', '月額', '健康', '厚生', '年金', '源泉', '所得税']
                for keyword in keywords:
                    if keyword in str(cell_value):
                        header_row = row
                        print(f"ヘッダー行発見: 行{row}")
                        break
            if header_row:
                break
        if header_row:
            break
    
    if header_row:
        print(f"\n=== 行{header_row}のヘッダー内容 ===")
        for col in range(1, 30):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                print(f"列{col} ({get_column_letter(col)}): {cell_value}")
        
        # データ行の確認（ヘッダーの次の行）
        print(f"\n=== 行{header_row + 1}のデータ内容 ===")
        for col in range(1, 30):
            cell_value = ws.cell(row=header_row + 1, column=col).value
            if cell_value is not None:
                print(f"列{col} ({get_column_letter(col)}): {cell_value}")
    
    # 渡邉俊行さんのデータを探す
    print(f"\n=== 渡邉俊行さんのデータ確認 ===")
    for row in range(1, 100):
        for col in range(1, 30):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str) and '渡邉俊行' in str(cell_value):
                print(f"渡邉俊行さん発見: 行{row}列{col} ({get_column_letter(col)}{row})")
                
                # その行の全データを表示
                print(f"行{row}の全データ:")
                for col_idx in range(1, 30):
                    cell_val = ws.cell(row=row, column=col_idx).value
                    if cell_val is not None:
                        print(f"  列{col_idx} ({get_column_letter(col_idx)}): {cell_val}")
                break
    
    wb.close()

if __name__ == "__main__":
    check_excel_headers() 