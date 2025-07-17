#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
標準報酬月額の値入力に関する詳細調査プログラム
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os

def debug_standard_salary_values():
    """標準報酬月額の値入力に関する詳細調査"""
    excel_file = "給与支給一覧令和7年.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"エラー: {excel_file} が見つかりません")
        return
    
    print("=== 標準報酬月額の値入力詳細調査 ===")
    print(f"ファイル: {excel_file}")
    print()
    
    # 通常の読み込み（数式あり）
    wb_normal = openpyxl.load_workbook(excel_file, data_only=False)
    ws_normal = wb_normal.active
    
    # データのみの読み込み（数式の計算結果）
    wb_data = openpyxl.load_workbook(excel_file, data_only=True)
    ws_data = wb_data.active
    
    # 標準報酬月額に関連するセルを探す
    print("=== 標準報酬月額に関連するセルの調査 ===")
    
    # ヘッダー行を探す
    header_row = None
    for row in range(1, 20):
        for col in range(1, 50):
            cell_value = ws_normal.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str):
                if '標準報酬月額' in str(cell_value) or '標準' in str(cell_value) and '月額' in str(cell_value):
                    header_row = row
                    print(f"標準報酬月額のヘッダー発見: 行{row}列{col} ({get_column_letter(col)}{row})")
                    print(f"  値: {cell_value}")
                    break
        if header_row:
            break
    
    if header_row:
        print(f"\nヘッダー行{header_row}の詳細:")
        for col in range(1, 30):
            cell_normal = ws_normal.cell(row=header_row, column=col)
            cell_data = ws_data.cell(row=header_row, column=col)
            cell_value = cell_normal.value
            if cell_value:
                print(f"列{col} ({get_column_letter(col)}): {cell_value}")
        
        # データ行の調査
        print(f"\n=== データ行の標準報酬月額調査 ===")
        for row in range(header_row + 1, header_row + 20):
            # 各列の値を確認
            for col in range(1, 30):
                cell_normal = ws_normal.cell(row=row, column=col)
                cell_data = ws_data.cell(row=row, column=col)
                
                cell_value_normal = cell_normal.value
                cell_value_data = cell_data.value
                cell_formula = cell_normal.value if cell_normal.data_type == 'f' else None
                
                # 数値データで、値が異なるセルを特定
                if (cell_value_normal is not None and 
                    cell_value_data is not None and 
                    cell_value_normal != cell_value_data):
                    
                    print(f"行{row}列{col} ({get_column_letter(col)}{row}):")
                    print(f"  通常読み込み値: {cell_value_normal} (型: {type(cell_value_normal)})")
                    print(f"  データ読み込み値: {cell_value_data} (型: {type(cell_value_data)})")
                    if cell_formula:
                        print(f"  数式: {cell_formula}")
                    print()
    
    # 特定の行（46行目以降）の標準報酬月額を詳しく調査
    print("=== 特定行の標準報酬月額詳細調査 ===")
    for row in range(46, 55):
        print(f"\n--- 行{row}の標準報酬月額 ---")
        
        # 標準報酬月額が含まれていそうな列を調査
        for col in range(1, 30):
            cell_normal = ws_normal.cell(row=row, column=col)
            cell_data = ws_data.cell(row=row, column=col)
            
            cell_value_normal = cell_normal.value
            cell_value_data = cell_data.value
            cell_formula = cell_normal.value if cell_normal.data_type == 'f' else None
            
            # 数値データで、値が異なるセルを特定
            if (cell_value_normal is not None and 
                cell_value_data is not None and 
                cell_value_normal != cell_value_data):
                
                print(f"列{col} ({get_column_letter(col)}):")
                print(f"  通常読み込み値: {cell_value_normal} (型: {type(cell_value_normal)})")
                print(f"  データ読み込み値: {cell_value_data} (型: {type(cell_value_data)})")
                if cell_formula:
                    print(f"  数式: {cell_formula}")
                print()
    
    # 空白セルの詳細調査
    print("=== 空白セルの詳細調査 ===")
    blank_cells = []
    for row in range(46, 55):
        for col in range(1, 30):
            cell_normal = ws_normal.cell(row=row, column=col)
            cell_data = ws_data.cell(row=row, column=col)
            
            # データ読み込みで空白だが、通常読み込みで値があるセル
            if (cell_data.value is None or cell_data.value == '') and cell_normal.value is not None:
                blank_cells.append((row, col, cell_normal.value))
    
    if blank_cells:
        print(f"データ読み込みで空白だが通常読み込みで値があるセル: {len(blank_cells)}個")
        for row, col, value in blank_cells:
            print(f"  行{row}列{col} ({get_column_letter(col)}{row}): {value}")
    else:
        print("データ読み込みで空白のセルは見つかりませんでした")
    
    # 現在のsalary_pdf_generator.pyで使用している列番号を確認
    print("\n=== 現在のプログラムでの列番号確認 ===")
    print("salary_pdf_generator.pyで標準報酬月額を取得している列番号を確認してください")
    
    wb_normal.close()
    wb_data.close()

if __name__ == "__main__":
    debug_standard_salary_values() 